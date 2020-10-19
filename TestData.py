#!/bin/python
# import http://docs.python-requests.org
import os
import openpyxl
import requests

class TestData():
    Status = "False"
    return_id = "NA"
    abs_path=os.path.abspath(os.getcwd())
    loc = (abs_path+"\\TestCases.xlsx")

    def __init__(self):
        self.data = []
        self.Jquery = {
            'key': '64da6c090937e433bc110eb78cf4da3f',
            'token': '16024eaa4e7af7480e7f44a436726f1791041c8e3227728447d9a8df7c3508f6',
                }
    def getidforData(self,Test_Precon):
        global objid
        wb_obj = openpyxl.load_workbook(self.loc)
        sheet_obj = wb_obj.active
        for row in range(2, sheet_obj.max_row):
            if(sheet_obj.cell(row,4).value==Test_Precon):
                objid=sheet_obj.cell(row,6).value
                break
    def DeleteBoard(self):
        url = "https://api.trello.com/1/boards/" + Boardid

        query = self.Jquery

        response = requests.request(
                "DELETE",
                url,params=query)

    def CreateandCompareBoard(self,BoardName):
            global Boardid , BoardName_get
            url = "https://api.trello.com/1/boards/"
            self.Jquery['name']=BoardName
            query=self.Jquery
            response = requests.request("POST",url,params=query)
            Boardid = ((response.text).split('"id":"')[1]).split('","')[0]
            self.return_id=Boardid
            print(Boardid)
            url = "https://api.trello.com/1/boards/" + Boardid
            headers = {
                "Accept": "application/json"
            }
            query=self.Jquery
            response = requests.request(
                "GET",
                url,
                headers=headers,
                params=query
            )
            BoardName_get = ((response.text).split('"name":"')[1]).split('","')[0]
            if (BoardName == BoardName_get):
                self.Status = "Pass"
            else:
                self.Status = "Fail"

    def CreateList(self,ListName):
        global Listid
        url = "https://api.trello.com/1/lists"
        self.Jquery['name'] = ListName
        self.Jquery['idBoard'] = Boardid
        query = self.Jquery
        response = requests.request("POST",url,params=query)
        Listid = ((response.text).split('"id":"')[1]).split('","')[0]
        self.return_id=Listid
        print(Listid)

    def CreateCard(self,CardName):
        #global Cardid
        url = "https://api.trello.com/1/cards"
        self.Jquery['idList'] = Listid
        self.Jquery['name'] = CardName
        query = self.Jquery
        response = requests.request("POST",url,params=query)
        Cardid = ((response.text).split('"id":"')[1]).split("\",\"")[0]
        self.return_id=Cardid
        self.GetCards(Cardid,CardName)
        if (CardName == CardName_get):
            self.Status = "Pass"
        else:
            self.Status = "Fail"

    def GetCards(self,Cardid,CardName):
        global CardName_get, CardID_get,CardText_get
        url = "https://api.trello.com/1/cards/"+Cardid
        headers = {
            "Accept": "application/json"
        }
        query = self.Jquery
        response = requests.request("GET",url,headers=headers,params=query)
        #remove this
        #print(json.dumps(json.loads(response.text), sort_keys=True, indent=4, separators=(",", ": ")))
        CardName_get = ((response.text).split('"name":"')[1]).split("\",\"")[0]
        CardID_get = ((response.text).split('"id":"')[1]).split("\",\"")[0]

    def EditCard(self,CardData):
        url = "https://api.trello.com/1/cards/"+objid
        headers = {
            "Accept": "application/json"
        }
        self.Jquery['name'] = CardData.split(',')[1]
        query = self.Jquery
        response = requests.request("PUT",url,headers=headers,params=query)
        self.GetCards(objid,CardData.split(',')[0])
        #print(json.dumps(json.loads(response.text), sort_keys=True, indent=4, separators=(",", ": ")))
        self.return_id=objid
        Cardid=objid
        if (CardData.split(',')[1] == CardName_get):
            self.Status = "Pass"
        else:
            self.Status = "Fail"

    def UpdateCard(self,CardData):
        url = "https://api.trello.com/1/cards/"+objid+"/actions/comments"
        headers = {
            "Accept": "application/json"
        }
        self.Jquery['text'] = CardData.split(',')[1]
        query = self.Jquery
        response = requests.request("POST",url,headers=headers,params=query)
        self.return_id = objid
        self.Status = "Pass"

    def DeleteCards(self,CardName):
        url = "https://api.trello.com/1/cards/"+objid
        query = self.Jquery
        response = requests.request("DELETE",        url,        params=query)
        self.return_id = objid
        self.Status = "Pass"


def main():
        objid=None
        tc = TestData()
        wb_obj = openpyxl.load_workbook(tc.loc)
        sheet_obj = wb_obj["API_TestCases"]
        for row in range(2, sheet_obj.max_row+1):
            Test_Step= sheet_obj.cell(row, 1).value
            Test_Key = sheet_obj.cell(row, 3).value
            Test_Data = sheet_obj.cell(row, 4).value
            Test_Precon = sheet_obj.cell(row, 5).value
            print("----------Test Step-----------" +Test_Step )
            try:
                if(Test_Precon != None):
                        tc.getidforData(Test_Precon)
            except:
                print("TestPreconError")
            funct ="tc."+Test_Key + "('" + Test_Data + "')"
            print("Function Name: "+ funct)
            exec(funct)
            print("Status: "+ tc.Status)
            sheet_obj.cell(row, 6).value=tc.return_id
            wb_obj.save(tc.loc)
            sheet_obj.cell(row, 7).value = tc.Status
            wb_obj.save(tc.loc)

#        tc.DeleteBoard()

if __name__ == '__main__':
    main()

