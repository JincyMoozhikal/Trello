#!/bin/python
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.wait import WebDriverWait
import os
import openpyxl
timeout = 10


class Driver():
    status = "fail"

    def __init__(self):
        self.driver = webdriver.Firefox()

    def waitload(self,driver,cssselector):
        try:
            element_present = EC.presence_of_element_located((By.CSS_SELECTOR, cssselector))
            WebDriverWait(driver, timeout).until(element_present)
        except TimeoutException:
            print("Timed out waiting for page to load")

    def waitload_ID(self,driver,id):
        try:
            element_present = EC.presence_of_element_located((By.ID, id))
            WebDriverWait(driver, timeout).until(element_present)
        except TimeoutException:
            print("Timed out waiting for page to load")


    def Openwebsite(self):
        try:
            self.driver.get("https://trello.com/")
            self.status="Pass"
        except:
            self.status = "Fail"

    def clickonLogin(self):
        try:
            self.waitload(self.driver,"div.float-right>a.text-white")
            self.driver.find_element_by_css_selector("div.float-right>a.text-white").click()
            self.waitload_ID(self.driver,'user')
            self.status="Pass"
        except:
            self.status = "Fail"

    def login(self):
        try:
            #enter login
            self.waitload_ID(self.driver,'user')
            self.driver.find_element_by_id("user").send_keys("jincy.moozhikal@gmail.com")
            self.driver.find_element_by_id("login").click()
            # enter pwd
            self.waitload_ID(self.driver, 'password')
            self.driver.find_element_by_id("password").send_keys("Test12345$")
            #submit form
            self.driver.find_element_by_id("login-submit").click()
            #wait for page to open
            self.waitload(self.driver,"li.boards-page-board-section-list-item>a.board-tile>div.board-tile-details>div.board-tile-details-name")
            self.driver.find_element_by_css_selector('li.boards-page-board-section-list-item>a.board-tile>div.board-tile-details>div.board-tile-details-name').click()
        except:
            self.driver.save_screenshot("screenshot.png")
            self.status = "Fail"

    def gotoboard(self):
        #click on boards
        try:
            self.waitload(self.driver,"div.card-composer-container>a.js-open-card-composer")
            self.driver.find_element_by_css_selector('div.card-composer-container>a.js-open-card-composer').click()
            self.status = "Pass"
        except:
            self.status = "Fail"

    def countBoards(self,count):
        try:
        #check elements
            elelist = []
            elelist = self.driver.find_elements(By.CLASS_NAME, 'js-card-details')
            TotalCards= len(elelist)
            if(TotalCards == count):
                self.status = "Pass"
            else:
                self.status = "Fail"
                self.driver.save_screenshot("screenshot.png")
        except:
            self.status = "Fail"

    def checkcommentoncard(self):
    #clicking on one with comment
        try:
            self.waitload(self.driver, "div.badge>span.icon-comment")
            if(self.driver.find_element_by_css_selector('div.badge>span.icon-comment').is_displayed()):
                self.status = "Pass"
            else:
                self.status = "Fail"
        except:
            self.status = "Fail"

    def addcommentoncard(self):
        try:
            self.driver.find_element_by_css_selector('div.badge>span.icon-comment').click()
    #add comment
            self.waitload(self.driver,"textarea.js-new-comment-input")
            self.driver.find_element_by_css_selector('textarea.js-new-comment-input').send_keys("New added comment")
    #save()
            self.driver.find_element_by_css_selector('div.comment-controls>input.js-add-comment').click()
    # close()
            self.waitload(self.driver, "div.window-wrapper>a.js-close-window")
            self.driver.find_element_by_css_selector('div.window-wrapper>a.js-close-window').click()
            self.status = "Pass"
        except:
            self.status = "Fail"



    def movecard(self):
    #move()
            try:
                fromA=self.driver.find_element_by_xpath('/html/body/div[1]/div[2]/div[1]/div[2]/main/div[3]/div/div[1]/div[3]/div[2]/div[1]/div/div[2]/a[2]/div[3]')
                to=self.driver.find_element_by_xpath('/html/body/div[1]/div[2]/div[1]/div[2]/main/div[3]/div/div[1]/div[3]/div[2]/div[4]/div/div[3]/a')
                action = ActionChains(self.driver)
                action.drag_and_drop(fromA,to)
                action.perform()
                self.status = "Pass"
            except:
                self.status = "Fail"
                print("Comment does not exist")

    def close(self):
        try:
            self.driver.close()
            self.status = "Pass"
        except:
            self.status = "Fail"


    #add another card
    #    waitload(driver,"span.js-add-another-card")
    #   driver.find_element_by_css_selector('span.js-add-another-card').click()

    # adding a text base here
    #  waitload(driver, "textarea.list-card-composer-textarea")
    # driver.find_element_by_css_selector('textarea.list-card-composer-textarea').send_keys("DDDDDDDDD")
    #submit the new card
    #driver.find_element_by_css_selector('input.primary').click()
def main():
        d=Driver()
        abs_path = os.path.abspath(os.getcwd())
        loc = (abs_path + "\\TestCases.xlsx")
        wb_obj = openpyxl.load_workbook(loc)
        sheet_obj = wb_obj["WebDriver_TestCases"]
        for row in range(2, sheet_obj.max_row+1):
            Test_Step = sheet_obj.cell(row, 1).value
            Test_Data = sheet_obj.cell(row, 4).value
            Test_Key = sheet_obj.cell(row, 3).value
            print("----------"+Test_Step+"-----------" )
            print("TestCase :  " + Test_Key)
            if(Test_Data ==  None):
                funct = "d." + Test_Key + "()"
            else:
                funct = "d." + Test_Key + "("+str(Test_Data)+")"
            try:
                exec(funct)
                print("Status :     " + d.status)
                sheet_obj.cell(row,5).value=d.status
            except:
                sheet_obj.cell(row, 5).value = "Fail"
            wb_obj.save(loc)




if __name__ == '__main__':
    main()
